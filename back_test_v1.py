import duckdb
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import configparser

# ========== 参数配置 ==========
MAX_HOLDING_TRADING_DAYS = 40   # 最大持有天数40天
INITIAL_CASH = 100000         # 每股初始购买金额10万元
BACKTEST_RESULT = {}
# ========== 参数配置 ==========

# 盈亏报告的字段映射
PROFIT_LOSS_MAPPING = {
    "no": "编号",
    "stock_code": "股票代码",
    "stock_name": "股票名称", 
    "init_cash": "初始金额",
    "bought_date": "购入日期",
    "total_shares": "总仓位",
    "cost_price": "成本价",
    "trade_date": "结算日期",
    "holding_days": "持有天数",
    "max_holding_days": "最大持有天数",
    "market_value": "市值(元)",
    "profit": "盈亏金额(元)",
    "profit_percent": "盈亏比"
}

# 字段顺序
PROFIT_LOSS_ORDER = [
    "no",
    "stock_code", 
    "stock_name", 
    "init_cash",
    "bought_date", 
    "total_shares",
    "cost_price",
    "trade_date",
    "holding_days",
    "max_holding_days",
    "market_value",
    "profit", 
    "profit_percent"
]

# 加载需要做回测运算的xlsx文件
def load_df_from_excel_file(file_path):
    df = None
    try:
        # 读取 Excel 文件的第一个工作表，第一行作为列名
        df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl', header=0)
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found")
    except Exception as e:
        print(f"Error: {str(e)}")
    return df

# 把df中某列的值转换为datetime格式
def convert_date_format_of_df_column(df, column_name="备注"):
    try:
        # 将“备注”列从 yyyyMMdd 转换为 yyyy-MM-dd
        df[column_name] = pd.to_datetime(df[column_name], format='%Y%m%d').dt.strftime('%Y-%m-%d')
        return df
    except Exception as e:
        print(f"Error converting dates in column '{column_name}': {str(e)}")
        return df

def load_target_df():
    df = load_df_from_excel_file("1009all.xlsx")
    convert_date_format_of_df_column(df=df)

    # 复制备注列为breakthrough_date
    df['stock_code'] = df['代码'].str.lower()
    df['stock_name'] = df['    名称']
    df['trade_date'] = df['备注']
    df['adj_stock_price'] = df['备注']
    return df

# 根据选中的突破日股票数据(结构:[{"stock_code": "AAPL"},{"stock_code": "TSM"}])
# 获取被选股票突破日后N天的交易数据
def get_next_N_days_data(stock_data_list, max_holding_days):
    """
    Connects to DuckDB, creates/ensures stock_data table exists (for testing),
    and queries stocks satisfying specific conditions using DuckDB.
    """

    # 创建 ConfigParser 对象
    config = configparser.ConfigParser()

    # 读取 .conf 文件
    config.read('./config.conf')
    earliest_time_limit=config['settings']['earliest_time_limit']                                   # 交易日期的最早时限，该日前的交易数据，不会被纳入选择
    cond1_and_cond3=config['settings']['cond1_and_cond3']                                           # 条件1和条件3的配置项。
    cond2=config['settings']['cond2']                                                               # 条件2：前N个交易日内有涨幅（大于等于5%）的K线
    apply_cond2_or_not=config['settings']['apply_cond2_or_not']                                     # 是否启用条件2：yes, 启用; no: 不启用。
    apply_cond5_or_not=config['settings']['apply_cond5_or_not']                                     # 是否启用条件5：yes, 启用; no: 不启用。
    # history_trading_days=config['settings']['history_trading_days']                               # 条件1：历史交易日选择范围。40: 40个交易日，60: 60个交易日，80: 80个交易日
    # main_board_amplitude_threshold=config['settings']['main_board_amplitude_threshold']           # 条件3：主板振幅。25: 25%, 30: 30%, 35: 35%
    # non_main_board_amplitude_threshold=config['settings']['non_main_board_amplitude_threshold']   # 条件3：创业板和科创板主板振幅。35: 35%， 40: 40%。
    history_trading_days=cond1_and_cond3.split('_')[0]
    main_board_amplitude_threshold=cond1_and_cond3.split('_')[1]
    non_main_board_amplitude_threshold=cond1_and_cond3.split('_')[2]
    max_market_capitalization=config['settings']['max_market_capitalization']                       # 最大流通市值，单位亿。
    min_market_capitalization=config['settings']['min_market_capitalization']                       # 最小流通市值，单位亿。
    net_profit_growth_rate=config['settings']['net_profit_growth_rate']                             # 净利润增长率。-20: -20%。
    total_revenue_growth_rate=config['settings']['total_revenue_growth_rate']                       # 营业总收入增长率。-20: -20%。
    use_cond_1_1_or_cond_1_2=config['settings']['use_cond_1_1_or_cond_1_2']                         # 使用条件1.1还是1.2进行筛选：1.1，使用条件1.1; 1.2, 使用条件1.2。
    range_days_of_cond_1_2=config['settings']['range_days_of_cond_1_2']                             # 使用条件1.2时，其后N个交易日设定值

    cond2_sql_where_clause = ''
    if apply_cond2_or_not == 'yes':
        cond2_sql_where_clause = 'AND has_gain_5_percent = 1'
    if apply_cond2_or_not == 'no':
        cond2_sql_where_clause = '-- AND has_gain_5_percent = 1'

    cond5_sql_where_clause = ''
    if apply_cond5_or_not == 'yes':
        cond5_sql_where_clause = f'AND net_profit_yoy >= {net_profit_growth_rate} AND revenue_yoy >= {total_revenue_growth_rate}'
    if apply_cond5_or_not == 'no':
        cond5_sql_where_clause = f''

    # Connect to DuckDB database file
    # Ensure 'stock_data.duckdb' exists and contains data,
    # or uncomment the data generation part below for testing.
    con = duckdb.connect(database='stock_data.duckdb', read_only=False)
    print("连接到数据库: stock_data.duckdb")
    
    stock_code_list = ", ".join(f"'{item['stock_code']}'" for item in stock_data_list)
    days_limit = 41 if max_holding_days is None else (int(max_holding_days) + 1)

    # Main Query SQL (optimized for DuckDB)
    # The SQL is mostly the same as DuckDB handles window functions efficiently.
    query_sql = f"""
    -- 📝 计算符合条件的股票交易日窗口
    WITH DeduplicatedStockData AS (
        -- ✅ 去掉 stock_data 中完全重复的行
        SELECT DISTINCT stock_code, stock_name, trade_date, open_price, close_price, high_price, low_price, prev_close_price, market_cap, total_market_cap, industry_level1, industry_level2, industry_level3 
        FROM stock_data
        -- 🔧 限定 stock_code 范围，只查询给定股票列表
        WHERE stock_code IN (
            -- ⚠️ 这里的 stock_code_list 可以是 Python 格式 ['AAPL','TSM'] 转换成 SQL 字符串 'AAPL','TSM'
            {stock_code_list}
        )
    ),
    StockWithRiseFall AS (
        -- ✅ 计算复权涨跌幅，公式: 复权涨跌幅 = 收盘价 / 前收盘价 - 1
        SELECT *,
            (close_price / NULLIF(prev_close_price, 0)) - 1 AS rise_fall
        FROM DeduplicatedStockData
    ),
    AdjustmentFactorComputed AS (
        -- ✅ 计算复权因子, 公式: 复权因子 = (1 + 复权涨跌幅).cumprod()
        SELECT *,
            EXP(SUM(LN(1 + rise_fall)) OVER (PARTITION BY stock_code ORDER BY trade_date)) AS adjustment_factor
        FROM StockWithRiseFall
    ),
    LastRecordComputed AS (
        -- ✅ 获取每个 stock_code 的最后一条记录的收盘价和复权因子
        SELECT 
            t.stock_code,
            t.close_price AS last_close_price,
            t.adjustment_factor AS last_adjustment_factor
        FROM (
            SELECT 
                stock_code,
                close_price,
                adjustment_factor,
                ROW_NUMBER() OVER (PARTITION BY stock_code ORDER BY trade_date DESC) AS rn
            FROM AdjustmentFactorComputed
        ) t
        WHERE t.rn = 1
    ),
    AdjustedStockData AS (
        SELECT 
            a.*,
            -- ✅ 计算前复权收盘价, 公式: 前复权收盘价 = 复权因子 * (最后一条数据的收盘价 / 最后一条数据的复权因子)
            a.adjustment_factor * (l.last_close_price / NULLIF(l.last_adjustment_factor, 0)) AS adj_close_price,
            -- ✅ 前复权其他价格
            (a.open_price / NULLIF(a.close_price, 0)) * (a.adjustment_factor * (l.last_close_price / NULLIF(l.last_adjustment_factor, 0))) AS adj_open_price,
            (a.high_price / NULLIF(a.close_price, 0)) * (a.adjustment_factor * (l.last_close_price / NULLIF(l.last_adjustment_factor, 0))) AS adj_high_price,
            (a.low_price / NULLIF(a.close_price, 0)) * (a.adjustment_factor * (l.last_close_price / NULLIF(l.last_adjustment_factor, 0))) AS adj_low_price,
            (a.prev_close_price / NULLIF(a.close_price, 0)) * (a.adjustment_factor * (l.last_close_price / NULLIF(l.last_adjustment_factor, 0))) AS adj_prev_close_price
        FROM AdjustmentFactorComputed a
        LEFT JOIN LastRecordComputed l ON a.stock_code = l.stock_code
    ),
    StockWindows AS (
        SELECT
            t.stock_code,
            t.trade_date,
            t.stock_name,
            t.close_price,
            t.high_price,
            t.low_price,
            t.open_price,
            t.adj_close_price,
            t.adj_high_price,
            t.adj_low_price,
            t.adj_open_price,
            t.industry_level1,
            t.industry_level2,
            t.industry_level3,
            -- ✅ 流通市值换算成“亿”
            (t.market_cap / 100000000) AS market_cap_of_100_million,
            -- 
            (t.total_market_cap / 100000000) AS total_market_cap_of_100_million,
            -- ✅ N个交易日内（不含当日）的最高收盘价, 使用的是复权后的收盘价
            MAX(t.adj_close_price) OVER (
                PARTITION BY t.stock_code
                ORDER BY t.trade_date
                ROWS BETWEEN {history_trading_days} PRECEDING AND 1 PRECEDING
            ) AS max_close_n_days,
            -- ✅ 对应的最高收盘价日期
            arg_max(t.trade_date, t.adj_close_price) OVER (
                PARTITION BY t.stock_code
                ORDER BY t.trade_date
                ROWS BETWEEN {history_trading_days} PRECEDING AND 1 PRECEDING
            ) AS max_close_n_days_date,
            -- ✅ N个交易日窗口内（不含当日）的最高价（用于振幅计算）, 使用的是复权后的最高价
            MAX(t.adj_high_price) OVER (
                PARTITION BY t.stock_code
                ORDER BY t.trade_date
                ROWS BETWEEN {history_trading_days} PRECEDING AND 1 PRECEDING
            ) AS max_high_n_days,
            -- ✅ N个交易日窗口内（不含当日）的最低价（用于振幅计算）, 使用的是复权后的最低价
            MIN(t.adj_low_price) OVER (
                PARTITION BY t.stock_code
                ORDER BY t.trade_date
                ROWS BETWEEN {history_trading_days} PRECEDING AND 1 PRECEDING
            ) AS min_low_n_days,
            -- ✅ N个交易日内（不含当日）的第一个交易日的开盘价，用作振幅分母。使用的是复权后的开盘价。
            FIRST_VALUE(t.adj_open_price) OVER (
                PARTITION BY t.stock_code
                ORDER BY t.trade_date
                ROWS BETWEEN {history_trading_days} PRECEDING AND 1 PRECEDING
            ) AS open_price_of_first_day_of_n_days,
            -- ✅ N个交易日内是否存在单日涨幅 ≥ 5%
            MAX(CASE
                WHEN (t.adj_close_price - t.adj_prev_close_price) / NULLIF(t.adj_prev_close_price, 0) >= {cond2} THEN 1
                ELSE 0
            END) OVER (
                PARTITION BY t.stock_code
                ORDER BY t.trade_date
                ROWS BETWEEN {history_trading_days} PRECEDING AND 1 PRECEDING
            ) AS has_gain_5_percent,
            -- ✅ 行号：确保窗口至少包含N个交易日
            ROW_NUMBER() OVER (
                PARTITION BY t.stock_code
                ORDER BY t.trade_date
            ) AS rn
        FROM
            AdjustedStockData t
        WHERE
            -- ✅ 排除北交所股票
            t.stock_code NOT LIKE 'bj%' AND
            -- ✅ 排除2022年1月1号之前的交易数据
            t.trade_date >= '{earliest_time_limit}'
    ),
    FilteredStockData AS (
        SELECT
            sw.stock_code,
            sw.stock_name,
            sw.trade_date,
            sw.adj_close_price,
            sw.adj_high_price,
            sw.adj_low_price,
            sw.adj_open_price,
            sw.max_close_n_days,
            sw.max_close_n_days_date,
            sw.market_cap_of_100_million,
            sw.total_market_cap_of_100_million,
            sw.industry_level1,
            sw.industry_level2,
            sw.industry_level3
        FROM
            StockWindows AS sw
        WHERE
            -- 📌 条件0：窗口内至少有N个交易日数据
            sw.rn > {history_trading_days}
            -- 📌 条件1：当日收盘价大于前N个交易日的最高收盘价的101%
            AND sw.adj_close_price > (sw.max_close_n_days * 1.01)
            -- 📌 条件2：前N个交易日内有涨幅（大于等于5%）的K线
            {cond2_sql_where_clause}
            -- 📌 条件3：前N个交易日的股票价格振幅度，上证和深证股票小于等于25%(30%, 35%)，创业板和科创板股票小于等于35%(40%, 40%)
            AND (
                -- ✅ 根据股票代码板块（前缀）确定振幅阈值
                CASE
                    WHEN sw.open_price_of_first_day_of_n_days > 0
                    THEN (sw.max_high_n_days - sw.min_low_n_days) * 1.0 / sw.open_price_of_first_day_of_n_days * 100
                    ELSE 999999 -- 避免除零错误
                END
            ) <= (
                CASE
                    -- ✅ 创业板（以300，301，302开头）或科创板（以688开头），小于等于35%(40%, 40%)
                    WHEN sw.stock_code LIKE 'sz300%' OR sw.stock_code LIKE 'sz301%' OR sw.stock_code LIKE 'sz302%' OR sw.stock_code LIKE 'sh688%' THEN {non_main_board_amplitude_threshold}
                    -- ✅ 上证主板（以600，601，603，605开头）小于等于25%(30%, 35%)
                    WHEN sw.stock_code LIKE 'sh600%' OR sw.stock_code LIKE 'sh601%' OR sw.stock_code LIKE 'sh603%' OR sw.stock_code LIKE 'sh605%' THEN {main_board_amplitude_threshold}
                    -- ✅ 深证主板（以000，001，002，003开头）小于等于25%(30%, 35%)
                    WHEN sw.stock_code LIKE 'sz000%' OR sw.stock_code LIKE 'sz001%' OR sw.stock_code LIKE 'sz002%' OR sw.stock_code LIKE 'sz003%' THEN {main_board_amplitude_threshold}
                    ELSE 1000
                END
            )
            -- 📌 条件4：流通市值在30亿至500亿之间
            AND sw.market_cap_of_100_million BETWEEN {min_market_capitalization} AND {max_market_capitalization}
    ),
    LimitedRangeStockData AS (
        -- 🔧 限定范围：每支股票从其 max_close_n_days_date 起，往后取 {days_limit} 个交易日数据
        SELECT *
        FROM StockWindows w
        WHERE EXISTS (
            SELECT 1
            FROM FilteredStockData f
            WHERE f.stock_code = w.stock_code
            AND w.trade_date BETWEEN f.max_close_n_days_date AND DATE_ADD(f.max_close_n_days_date, INTERVAL {days_limit} DAY)
        )
    )
    -- ✅ 最终输出
    SELECT
        stock_code,
        stock_name,
        trade_date,
        max_close_n_days_date AS breakthrough_date,
        ROUND(max_close_n_days, 2) AS adj_support_price,
        ROUND(adj_close_price, 2) AS close,
        ROUND(adj_high_price, 2) AS high,
        ROUND(adj_low_price, 2) AS low,
        ROUND(adj_open_price, 2) AS open,
        industry_level2,
        industry_level3
    FROM LimitedRangeStockData
    ORDER BY stock_code, trade_date;
    """
    
    # 获取查询结果
    results_df = con.execute(query_sql).fetchdf()
    
    # 关闭连接
    con.close()

    #返回查询结果
    return results_df


def update_position(stock_code, stock_name, support_date, support_price, trade_type, trade_date, trade_positions, trade_price, close_price, holding_days):
    global BACKTEST_RESULT, INITIAL_CASH, MAX_HOLDING_TRADING_DAYS
    if stock_code in BACKTEST_RESULT:
        stock_data = BACKTEST_RESULT[stock_code]
        stock_data["stock_code"] = stock_code
        stock_data["stock_name"] = stock_name
        stock_data["init_cash"] = INITIAL_CASH
        stock_data["support_date"] = support_date
        stock_data["support_price"] = support_price
        stock_data["max_holding_days"] = MAX_HOLDING_TRADING_DAYS
        stock_data["holding_days"] = holding_days
        if trade_type == "sell":
            stock_data["trade_date"] = max(trade_date, stock_data["trade_date"]) if "trade_date" in stock_data else trade_date
            stock_data["current_positions"] = stock_data["current_positions"] - trade_positions
            stock_data["current_cash"] = stock_data["current_cash"] + (trade_price * trade_positions)
            stock_data["market_value"] = stock_data["current_positions"] * close_price
            stock_data["profit"] =  stock_data["market_value"] + stock_data["current_cash"] - INITIAL_CASH
            stock_data["profit_percent"] = ((stock_data["profit"] / INITIAL_CASH) * 100)
        if trade_type == "buy":
            stock_data["bought_date"] = min(trade_date, stock_data["bought_date"])
            stock_data["total_shares"] = stock_data["total_shares"] + trade_positions
            stock_data["current_positions"] = stock_data["current_positions"] + trade_positions
            stock_data["current_cash"] = stock_data["current_cash"] - (trade_price * trade_positions)
            stock_data["cost_price"] = ((INITIAL_CASH - stock_data["current_cash"]) / stock_data["current_positions"])
            stock_data["market_value"] = stock_data["current_positions"] * close_price
            stock_data["profit"] =  stock_data["market_value"] + stock_data["current_cash"] - INITIAL_CASH
            stock_data["profit_percent"] = ((stock_data["profit"] / INITIAL_CASH) * 100)
    else:
        if trade_type == "buy":
            stock_data = {}
            stock_data["stock_code"] = stock_code
            stock_data["stock_name"] = stock_name
            stock_data["init_cash"] = INITIAL_CASH
            stock_data["support_date"] = support_date
            stock_data["support_price"] = support_price
            stock_data["max_holding_days"] = MAX_HOLDING_TRADING_DAYS
            stock_data["holding_days"] = holding_days
            stock_data["bought_date"] = trade_date
            stock_data["cost_price"] = trade_price
            stock_data["current_cash"] = INITIAL_CASH - (trade_price * trade_positions)
            stock_data["current_positions"] = trade_positions
            stock_data["total_shares"] = trade_positions
            stock_data["market_value"] = stock_data["current_positions"] * close_price
            stock_data["profit"] = 0.00
            stock_data["profit_percent"] = 0.00
            BACKTEST_RESULT[stock_code] = stock_data

def do_back_test():
    global BACKTEST_RESULT
    BACKTEST_RESULT.clear()
    # 获取数据
    target_df = load_target_df()
    stock_data_list = target_df[['stock_code', 'stock_name']].to_dict('records')
    stock_df = get_next_N_days_data(stock_data_list, MAX_HOLDING_TRADING_DAYS)

    # 转换日期类型
    stock_df['trade_date'] = pd.to_datetime(stock_df['trade_date'])
    stock_df['breakthrough_date'] = pd.to_datetime(stock_df['breakthrough_date'])

    # 按stock_code和trade_date进行排序
    stock_df = stock_df.sort_values(['stock_code', 'trade_date'])

    # results = []
    # stock_to_remaining = {}

    # 循环处理每一支要回测的股票数据
    for idx, target in target_df.iterrows():
        stock_code = target['stock_code']
        stock_name = target['stock_name']
        
        group = stock_df[stock_df['stock_code'] == stock_code].reset_index(drop=True)
        if group.empty:
            continue

        support_date = stock_df.loc[stock_df['stock_code'] == stock_code, 'breakthrough_date'].iloc[0]
        support_price = stock_df.loc[stock_df['stock_code'] == stock_code, 'adj_support_price'].iloc[0]
        
        # 找到突破日的后一日
        next_days = group[group['trade_date'] > support_date]
        if next_days.empty:
            continue
        bought_date = next_days['trade_date'].iloc[0]
        bought_idx = group[group['trade_date'] == bought_date].index[0]
        
        # 买入策略：以开盘价买入50%， 以收盘价买入50%。按100的整数倍仓位进行购买，剩余按现金进行持有。
        initial_cash = INITIAL_CASH
        cash_morning = initial_cash * 0.5
        cash_evening = initial_cash * 0.5
        
        shares_morning = ((cash_morning / group.loc[bought_idx, 'open']) // 100) * 100
        cost_morning = shares_morning * group.loc[bought_idx, 'open']
        # remaining_morning = cash_morning - cost_morning
        update_position(
            stock_code, stock_name, support_date, support_price, "buy", bought_date,
            shares_morning, group.loc[bought_idx, 'open'], group.loc[bought_idx, 'close'], 1
        )
        
        shares_evening = ((cash_evening / group.loc[bought_idx, 'close']) // 100) * 100
        cost_evening = shares_evening * group.loc[bought_idx, 'close']
        # remaining_evening = cash_evening - cost_evening
        update_position(
            stock_code, stock_name, support_date, support_price, "buy", bought_date,
            shares_evening, group.loc[bought_idx, 'close'], group.loc[bought_idx, 'close'], 1
        )
        
        total_shares = shares_morning + shares_evening
        if total_shares == 0:
            print(f"初始资金不足以买入至少100股，忽略对股票: {stock_name}({stock_code}) 进行回测。")
            continue    # 如果购入仓位为0，则忽略该股票。
        
        total_cost = cost_morning + cost_evening
        # remaining_cash = remaining_morning + remaining_evening
        cost_price = total_cost / total_shares
        
        # stock_to_remaining[stock_code] = remaining_cash
        
        # 初始仓位
        current_position = total_shares
        stop_loss = cost_price * 0.95
        half_sold = False
        recover_count = 0
        holding = True
        
        # 卖出策略（设条件单）
        max_rise = 1.0
        rise_break_date = None
        rise_count = 0
        
        # 初始化交易日计数
        holding_days = 0
        
        for i in range(bought_idx, len(group)):
            if not holding:
                break
            
            # 增加交易日计数
            holding_days += 1
            current_date = group.loc[i, 'trade_date']
            current_open = group.loc[i, 'open']
            current_high = group.loc[i, 'high']
            current_low = group.loc[i, 'low']
            current_close = group.loc[i, 'close']
            
            # 如果持有天数超过40个交易日，直接卖出（用收盘卖）
            if holding_days > MAX_HOLDING_TRADING_DAYS:
                if i > bought_idx:
                    # 在第40个交易日卖出
                    prev_i = i - 1
                    sell_price = group.loc[prev_i, 'close']
                    current_date = group.loc[prev_i, 'trade_date']
                    holding_days -= 1  # 回退到前一天的交易日计数
                    update_position(
                        stock_code, stock_name, support_date, support_price, "sell", current_date,
                        current_position, sell_price, current_close, holding_days
                    )
                    holding = False
                    continue
            
            # 检查止损价，低于止损价就卖出（用止损价卖）
            if current_low < stop_loss:
                sell_price = stop_loss
                update_position(
                    stock_code, stock_name, support_date, support_price, "sell", current_date,
                    current_position, sell_price, current_close, holding_days
                )
                holding = False
                continue
            
            # 跌破支撑线但未跌破止损线，3日收不上去清仓。（按收盘价卖）
            if current_low < support_price:
                if current_close >= support_price:
                    recover_count = 0
                else:
                    recover_count += 1
                if recover_count >= 4:
                    sell_price = current_close
                    update_position(
                        stock_code, stock_name, support_date, support_price, "sell", current_date,
                        current_position, sell_price, current_close, holding_days
                    )
                    holding = False
                    continue
            else:
                recover_count = 0
            
            # Current rise
            current_rise = current_high / cost_price
            
            # 涨幅达到10%时，卖出50%仓位（用的是成本价*1.1卖出）
            if not half_sold and current_rise >= 1.10:
                sell_position = current_position * 0.5
                current_position -= sell_position
                sell_price = cost_price * 1.10
                update_position(
                    stock_code, stock_name, support_date, support_price, "sell", current_date,
                    sell_position, sell_price, current_close, holding_days
                )
                half_sold = True
                stop_loss = cost_price * 1.10
                max_rise = 1.10
            
            # 剩余 50% 仓位的动态跟踪策略
            if half_sold:
                if current_rise > max_rise:
                    if max_rise < 1.20 and current_rise >= 1.20:
                        rise_break_date = current_date
                        rise_count = 0
                        max_rise = 1.20
                        stop_loss = cost_price * 1.20
                    elif max_rise < 1.30 and current_rise >= 1.30:
                        rise_break_date = current_date
                        rise_count = 0
                        max_rise = 1.30
                        stop_loss = cost_price * 1.30
                    elif max_rise < 1.40 and current_rise >= 1.40:
                        rise_break_date = current_date
                        rise_count = 0
                        max_rise = 1.40
                        stop_loss = cost_price * 1.40
                    elif max_rise < 1.50 and current_rise >= 1.50:
                        rise_break_date = current_date
                        rise_count = 0
                        max_rise = 1.50
                        stop_loss = cost_price * 1.50
                    elif max_rise < 1.60 and current_rise >= 1.60:
                        rise_break_date = current_date
                        rise_count = 0
                        max_rise = 1.60
                        stop_loss = cost_price * 1.60
                    elif max_rise < 1.70 and current_rise >= 1.70:
                        rise_break_date = current_date
                        rise_count = 0
                        max_rise = 1.70
                        stop_loss = cost_price * 1.70
                    elif max_rise < 1.80 and current_rise >= 1.80:
                        rise_break_date = current_date
                        rise_count = 0
                        max_rise = 1.80
                        stop_loss = cost_price * 1.80
                    elif max_rise < 1.90 and current_rise >= 1.90:
                        rise_break_date = current_date
                        rise_count = 0
                        max_rise = 1.90
                        stop_loss = cost_price * 1.90
                    
                    # 最高到200%，届时止损线不再调整，直接清仓。（按200%价格卖）
                    if current_rise >= 2.00:
                        sell_price = cost_price * 2.00
                        update_position(
                            stock_code, stock_name, support_date, support_price, "sell", current_date,
                            current_position, sell_price, current_close, holding_days
                        )
                        current_position = 0
                        holding = False
                        continue
                
                # 2.2 否则若突破130%后，5个交易日不超过140%清仓。（按收盘价卖）
                # if rise_break_date is not None:
                #     rise_count += 1  # 交易日计数
                #     next_level = max_rise + 0.10
                #     if rise_count >= 5 and current_rise < next_level:
                #         sell_price = current_close
                #         update_position(
                #             stock_code, stock_name, support_date, support_price, "sell", current_date,
                #             current_position, sell_price, current_close, holding_days
                #         )
                #         current_position = 0
                #         holding = False
                #         continue
                
                # 1. 回调至130%，立即卖出；（按130%价卖）
                if current_close < stop_loss:
                    sell_price = stop_loss
                    update_position(
                        stock_code, stock_name, support_date, support_price, "sell", current_date,
                        current_position, sell_price, current_close, holding_days
                    )
                    current_position = 0
                    holding = False
                    continue
                
                # 2.1 若未回调，但持有满40天，当天收盘前卖出。（按收盘价卖）
                if max_rise >= current_rise and holding_days >= 40:
                    sell_price = current_close
                    update_position(
                        stock_code, stock_name, support_date, support_price, "sell", current_date,
                        current_position, sell_price, current_close, holding_days
                    )
                    current_position = 0
                    holding = False
                    continue
            
            # 最多持有40天，无论多少清仓（按收盘价卖）
            if i == len(group) - 1 and holding:
                sell_price = current_close
                update_position(
                    stock_code, stock_name, support_date, support_price, "sell", current_date,
                    current_position, sell_price, current_close, holding_days
                )
                current_position = 0
                holding = False

    # 以下部分保持不变
    final_df = pd.DataFrame(list(BACKTEST_RESULT.values()))

    # 按股票代码和股票名称对交易数据进行汇总
    merged_df = final_df.groupby(['stock_code', 'stock_name']).agg(
        bought_date=('bought_date', 'min'),
        init_cash=('init_cash', 'max'),
        total_shares=('total_shares', 'max'),
        cost_price=('cost_price', 'max'),
        trade_date=('trade_date', 'max'),
        support_date=('support_date', 'max'),
        support_price=('support_price', 'max'),
        current_cash=('current_cash', 'sum'),
        holding_days=('holding_days', 'max'),
        max_holding_days=('max_holding_days', 'max'),
        market_value=('market_value', 'sum'),
        profit=('profit', 'sum')
    ).reset_index()

    # 如果有剩余现金，把剩余现金计入账户市值
    for idx, row in merged_df.iterrows():
        stock_code = row['stock_code']
        current_cash = row['current_cash']
        merged_df.at[idx, 'market_value'] += current_cash

    # 根据账户市值和初始资金计算利润和利润率
    merged_df['profit'] = merged_df['market_value'] - merged_df['init_cash']
    merged_df['profit_percent'] = ( merged_df['profit'] / merged_df['init_cash']).round(2)

    # 添加编号列
    merged_df['no'] = range(1, len(merged_df) + 1)

    # 重命名列为中文
    merged_df = merged_df.rename(columns=PROFIT_LOSS_MAPPING)
    # 按指定顺序重新排列列
    merged_df = merged_df[[PROFIT_LOSS_MAPPING[col] for col in PROFIT_LOSS_ORDER]]

    total_init_cash = merged_df[PROFIT_LOSS_MAPPING['init_cash']].sum()
    total_market_value = merged_df[PROFIT_LOSS_MAPPING['market_value']].sum()
    total_profit_percent = ((total_market_value - total_init_cash) / total_init_cash).round(2)
    
    # 构造汇总行，使用中文列名
    total_row = pd.DataFrame({
        PROFIT_LOSS_MAPPING['no']: [None],
        PROFIT_LOSS_MAPPING['stock_code']: [None],
        PROFIT_LOSS_MAPPING['stock_name']: ['总投资金额'],
        PROFIT_LOSS_MAPPING['init_cash']: [total_init_cash],
        PROFIT_LOSS_MAPPING['bought_date']: [None],
        PROFIT_LOSS_MAPPING['total_shares']: [None],
        PROFIT_LOSS_MAPPING['cost_price']: [None],
        PROFIT_LOSS_MAPPING['trade_date']: [None],
        PROFIT_LOSS_MAPPING['holding_days']: [None],
        PROFIT_LOSS_MAPPING['max_holding_days']: ['总市值'],
        PROFIT_LOSS_MAPPING['market_value']: [total_market_value],
        PROFIT_LOSS_MAPPING['profit']: [total_market_value - total_init_cash],
        PROFIT_LOSS_MAPPING['profit_percent']: [total_profit_percent]
    })

    # 准备导出数据
    final_export_df = pd.concat([merged_df, total_row], ignore_index=True)

    # 设置导出回测结果文件名称
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"组合盈亏报告_{timestamp}.xlsx"

    # 导出回测结果文件到excel
    wb = Workbook()
    ws = wb.active
    ws.title = "组合盈亏报告"

    for r in dataframe_to_rows(final_export_df, index=False, header=True):
        ws.append(r)

    # 红的行表示盈利、绿的行表示亏损
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for row in range(2, len(merged_df) + 2):
        profit_pct = final_export_df.loc[row-2, PROFIT_LOSS_MAPPING['profit_percent']]
        if profit_pct > 0:
            for col in range(1, len(final_export_df.columns) + 1):
                ws.cell(row=row, column=col).fill = red_fill
        else:
            for col in range(1, len(final_export_df.columns) + 1):
                ws.cell(row=row, column=col).fill = green_fill

    col_map = {col: idx+1 for idx, col in enumerate(final_export_df.columns)}
    for row in range(2, ws.max_row + 1):
        for date_col in [PROFIT_LOSS_MAPPING['bought_date'], PROFIT_LOSS_MAPPING['trade_date']]:
            if date_col in col_map:
                cell = ws.cell(row, col_map[date_col])
                if cell.value is not None:
                    cell.number_format = 'yyyy-mm-dd'
        
        for float_col in [PROFIT_LOSS_MAPPING['cost_price'], PROFIT_LOSS_MAPPING['market_value'], PROFIT_LOSS_MAPPING['profit']]:
            if float_col in col_map:
                cell = ws.cell(row, col_map[float_col])
                if cell.value is not None:
                    cell.number_format = '0.00'
        
        if PROFIT_LOSS_MAPPING['profit_percent'] in col_map:
            cell = ws.cell(row, col_map[PROFIT_LOSS_MAPPING['profit_percent']])
            if cell.value is not None:
                cell.number_format = '0.00%'
        
        for int_col in [PROFIT_LOSS_MAPPING['init_cash'], PROFIT_LOSS_MAPPING['total_shares'], PROFIT_LOSS_MAPPING['holding_days'], PROFIT_LOSS_MAPPING['max_holding_days']]:
            if int_col in col_map:
                cell = ws.cell(row, col_map[int_col])
                if cell.value is not None:
                    cell.number_format = '0'

    wb.save(filename)

if __name__ == '__main__':
    do_back_test()