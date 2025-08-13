import backtrader as bt
import pandas as pd
import duckdb
import numpy as np
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import configparser
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment, Font

# 盈亏报告的字段映射
PROFIT_LOSS_MAPPING = {
    "no": "编号",
    "stock_code": "股票代码",
    "stock_name": "股票名称", 
    "bought_date": "购入日期",
    "init_cash": "初始金额",
    "trade_date": "结算日期",
    "holding_days": "持有天数",
    "market_value": "市值(元)",
    "profit": "盈亏金额(元)",
    "profit_percent": "盈亏比"
}

# 字段顺序
PROFIT_LOSS_ORDER = [
    "no",
    "stock_code", 
    "stock_name", 
    "bought_date", 
    "init_cash",
    "holding_days", 
    "trade_date",
    "market_value",
    "profit", 
    "profit_percent"
]

# 回测结果
PROFIT_AND_LOSS_SITUATION = []

# 初始金额
INITIAL_CASH = 100000

# 根据选中的突破日股票数据(结构:[{"trade_date": '2025-07-01', "stock_code": "AAPL"},{"trade_date": '2025-08-04', "stock_code": "AAPL"}])
# 获取被选股票突破日后N天的交易数据
def get_next_N_days_data(stock_data_list, holding_day):
    # 连接到DuckDB数据库
    con = duckdb.connect('stock_data.duckdb')
    
    # 初始化一个空的DataFrame，用于存储所有股票的查询结果
    all_results = []
    
    # 遍历所有的突破日股票记录
    for record in stock_data_list:
        stock_code = record['stock_code']
        trade_date = record['trade_date']
        
        # SQL查询：从突破日开始查询后20个交易日的数据
        query = f"""
        -- 📝 计算符合条件的股票交易日窗口
        WITH DeduplicatedStockData AS (
            -- ✅ 去掉 stock_data 中完全重复的行
            SELECT DISTINCT stock_code, stock_name, trade_date, open_price, close_price, high_price, low_price, prev_close_price, market_cap, industry_level2, industry_level3, volume FROM stock_data
        ),
        StockWithRiseFall AS (
            -- ✅ 计算复权涨跌幅，公式: 复权涨跌幅 = 收盘价 / 前收盘价 - 1
            SELECT *,
                (close_price / NULLIF(prev_close_price, 0)) - 1 AS rise_fall
            FROM DeduplicatedStockData
        ),
        AdjustmentFactorComputed AS (
            -- ✅ 计算复权因子, 公式: 复权因子 = (1 + 复权涨跌幅).cumprod()
            SELECT *,
                EXP(SUM(LN(1 + rise_fall)) OVER (PARTITION BY stock_code ORDER BY trade_date)) AS adjustment_factor
            FROM StockWithRiseFall
        ),
        LastRecordComputed AS (
            -- ✅ 获取每个 stock_code 的最后一条记录的收盘价和复权因子
            SELECT 
                t.stock_code,
                t.close_price AS last_close_price,
                t.adjustment_factor AS last_adjustment_factor
            FROM (
                SELECT 
                    stock_code,
                    close_price,
                    adjustment_factor,
                    ROW_NUMBER() OVER (PARTITION BY stock_code ORDER BY trade_date DESC) AS rn
                FROM AdjustmentFactorComputed
            ) t
            WHERE t.rn = 1
        ),
        AdjustedStockData AS (
            SELECT 
                a.*,
                -- ✅ 计算前复权收盘价, 公式: 前复权收盘价 = 复权因子 * (最后一条数据的收盘价 / 最后一条数据的复权因子)
                a.adjustment_factor * (l.last_close_price / NULLIF(l.last_adjustment_factor, 0)) AS adj_close_price,
                -- ✅ 前复权其他价格
                (a.open_price / NULLIF(a.close_price, 0)) * (a.adjustment_factor * (l.last_close_price / NULLIF(l.last_adjustment_factor, 0))) AS adj_open_price,
                (a.high_price / NULLIF(a.close_price, 0)) * (a.adjustment_factor * (l.last_close_price / NULLIF(l.last_adjustment_factor, 0))) AS adj_high_price,
                (a.low_price / NULLIF(a.close_price, 0)) * (a.adjustment_factor * (l.last_close_price / NULLIF(l.last_adjustment_factor, 0))) AS adj_low_price,
                (a.prev_close_price / NULLIF(a.close_price, 0)) * (a.adjustment_factor * (l.last_close_price / NULLIF(l.last_adjustment_factor, 0))) AS adj_prev_close_price
            FROM AdjustmentFactorComputed a
            LEFT JOIN LastRecordComputed l ON a.stock_code = l.stock_code
        ),
        RankedData AS (
            SELECT
                t.stock_code,
                t.trade_date,
                t.stock_name,
                t.volume,
                t.adj_close_price,
                t.adj_high_price,
                t.adj_low_price,
                t.adj_open_price,
                t.industry_level2,
                t.industry_level3,
                row_number() OVER (PARTITION BY t.stock_code ORDER BY t.trade_date) AS rn
            FROM AdjustedStockData t
            WHERE t.stock_code = '{stock_code}' AND t.trade_date >= '{trade_date}'
        )
        SELECT
            stock_code,
            trade_date,
            stock_name,
            volume,
            adj_close_price AS close,
            adj_high_price AS high,
            adj_low_price AS low,
            adj_open_price AS open,
            industry_level2,
            industry_level3
        FROM RankedData
        WHERE rn > (SELECT row_number() OVER (PARTITION BY stock_code ORDER BY trade_date) FROM RankedData WHERE trade_date = '{trade_date}' LIMIT 1)
        AND rn <= (SELECT row_number() OVER (PARTITION BY stock_code ORDER BY trade_date) FROM RankedData WHERE trade_date = '{trade_date}' LIMIT 1) + {holding_day} + 1
        ORDER BY trade_date;
        """

        # 执行查询并获取结果
        df = con.execute(query).fetchdf()
        
        # 将查询结果添加到 all_results 列表中
        all_results.append(df)
    
    # 将所有结果合并成一个大的DataFrame
    final_df = pd.concat(all_results, ignore_index=True)
    con.close()  # 关闭数据库连接
    return final_df

# 买入策略
class MyStrategy(bt.Strategy):
    # 设置参数，支持传入突破日和支撑位信息
    params = (
        ('breakout_date', None),            # 突破日
        ('first_support_price', None),      # 第一支撑位价格
    )

    def __init__(self):
        # 初始化所需的变量和字段
        self.breakout_date = self.params.breakout_date
        self.first_support_price = self.params.first_support_price

        self.buy_stage = {}  # 使用字典跟踪每只股票的买入阶段
        self.orders = {}  # 跟踪订单状态

        # 为每只数据初始化买入阶段
        for data in self.datas:
            # 增加一个变量记录买入阶段：0未买入，1买入50%，2买入全部
            self.buy_stage[data._name] = 0
            self.orders[data._name] = None
    
    def notify_order(self, order):
        """订单状态通知"""
        stock_code = order.data._name
        if order.status in [order.Completed]:
            if order.isbuy():
                print(f"买入订单完成：股票 {stock_code}, 数量: {order.executed.size}, " f"价格: {order.executed.price:.2f}, 日期: {self.datas[0].datetime.date(0)}")
            elif order.issell():
                print(f"卖出订单完成：股票 {stock_code}, 数量: {order.executed.size}, " f"价格: {order.executed.price:.2f}, 日期: {self.datas[0].datetime.date(0)}")
            self.orders[stock_code] = None  # 清除订单
        # elif order.status in [order.Submitted, order.Accepted]:
        #     print(f"订单已提交/接受: 股票 {stock_code}, 价格={order.price:.2f}, 数量={order.size}")
        elif order.status in [order.Canceled, order.Margin, order.Expired, order.Rejected]:
            print(f"订单失败：股票 {stock_code}, 状态: {order.Status[order.status]}")
            self.orders[stock_code] = None

    def next(self):
        for data in self.datas:
            stock_code = data._name
            adj_close_price = round(data.close[0], 2)
            adj_open_price = round(data.open[0], 2)
            adj_low_price = round(data.low[0], 2)
            adj_high_price = round(data.high[0], 2)
            current_date = data.datetime.date(0)

            # 跳过无效价格
            if not adj_close_price or not adj_open_price or adj_close_price <= 0 or adj_open_price <= 0:
                print(f"股票 {stock_code} 在 {current_date} 的价格无效，跳过")
                continue

            # 如果有未完成的订单，跳过
            if self.orders.get(stock_code):
                continue

            if self.buy_stage[stock_code] == 0:
                cash_to_use = self.broker.get_cash() * 0.5
                remaining_cash = self.broker.get_cash() - cash_to_use - self.broker.get_fundvalue()

                total_buy_size = 0
                average_price = 0.00
                total_cost = 0.00

                first_buy_size = int(cash_to_use / adj_open_price)
                first_buy_price = 0.00

                if first_buy_size > 0:
                    # self.orders[stock_code] = self.buy(data=data, size=first_buy_size, price=adj_open_price)
                    self.buy_stage[stock_code] = 1
                    first_buy_price = adj_open_price
                    print(f"开盘买入50%：股票 {stock_code}, 数量: {first_buy_size}, 价格: {adj_open_price:.2f}")

                # 增加 1 天
                next_day = current_date + timedelta(days=1)

                second_buy_size = 0
                second_buy_price = 0.00
                # 回踩支撑位买入剩余50%
                if self.buy_stage[stock_code] == 1 and adj_low_price <= self.first_support_price and self.first_support_price <= adj_high_price:
                    second_buy_size = int(remaining_cash / self.first_support_price)
                    second_buy_price = self.first_support_price

                    total_buy_size = first_buy_size + second_buy_size
                    average_price = round(((first_buy_size * first_buy_price + second_buy_size * second_buy_price) / total_buy_size), 2)
                    total_cost = round((total_buy_size * average_price), 2)
                    while total_cost > (self.broker.get_cash() - self.broker.get_fundvalue()):
                        second_buy_size = second_buy_size - 1
                        total_buy_size = first_buy_size + second_buy_size
                        average_price = round(((first_buy_size * first_buy_price + second_buy_size * second_buy_price) / total_buy_size), 2)
                        total_cost = round((total_buy_size * average_price), 2)
                    
                    if second_buy_size > 0:
                        # self.orders[stock_code] = self.buy(data=data, size=second_buy_size, price=self.first_support_price)
                        # self.orders[stock_code] = self.buy(data=data, size=second_buy_size, price=self.first_support_price, exectype=bt.Order.Limit, valid=next_day)
                        # self.buy_stage[stock_code] = 2
                        print(f"回踩第一支撑位买入剩余50%：股票 {stock_code}, 数量: {second_buy_size}, 价格: {self.first_support_price:.2f}")

                # 按收盘价买入剩余50%
                if self.buy_stage[stock_code] == 1 and adj_low_price > self.first_support_price:
                    second_buy_size = int(remaining_cash / adj_close_price)
                    second_buy_price = adj_close_price

                    total_buy_size = first_buy_size + second_buy_size
                    average_price = round(((first_buy_size * first_buy_price + second_buy_size * second_buy_price) / total_buy_size), 2)
                    total_cost = round((total_buy_size * average_price), 2)
                    while total_cost > (self.broker.get_cash() - self.broker.get_fundvalue()):
                        second_buy_size = second_buy_size - 1
                        total_buy_size = first_buy_size + second_buy_size
                        average_price = round(((first_buy_size * first_buy_price + second_buy_size * second_buy_price) / total_buy_size), 2)
                        total_cost = round((total_buy_size * average_price), 2)

                    if second_buy_size > 0:
                        # self.orders[stock_code] = self.buy(data=data, size=second_buy_size, price=adj_close_price)
                        # self.orders[stock_code] = self.buy(data=data, size=second_buy_size, price=adj_close_price, exectype=bt.Order.Limit, valid=next_day)
                        # self.buy_stage[stock_code] = 2
                        print(f"按收盘价买入剩余50%：股票 {stock_code}, 数量: {second_buy_size}, 价格: {adj_close_price:.2f}")
                
                # 因为基本框架不支持一天内使用两个价格进行买入，则干脆使用均价进行一次买入来进行模拟
                print(f"总持仓: {total_buy_size}, 均价: {average_price:.2f}, 日最高价: {adj_high_price:.2f}, 日最低价: {adj_low_price:.2f} 总成本: {total_cost:.2f}")
                self.orders[stock_code] = self.buy(data=data, size=total_buy_size, price=average_price, exectype=bt.Order.Limit, valid=next_day)
                # self.orders[stock_code] = self.buy(data=data, size=total_buy_size, price=average_price)
                self.buy_stage[stock_code] = 2

            profit = self.broker.get_value() - self.broker.startingcash
            profit_percent = profit / self.broker.startingcash * 100
            print(f"股票: {stock_code}, 日期: {current_date}, 收盘价: {adj_close_price:.2f}, " f"开盘价: {adj_open_price:.2f}, 持仓: {self.getposition(data).size}, 均价: {self.getposition(data).price:.2f}, 盈亏: {profit:.2f}, 盈亏比: {profit_percent:.2f}")

            # 加入退出策略：如果亏损超过3%，则直接平仓退出
            if profit_percent <= -3 and self.getposition(data).size > 0:
                self.close()

    def stop(self):
        # 计算最终收益
        for data in self.datas:
            profit = self.broker.get_value() - self.broker.startingcash
            profit_per = profit / self.broker.startingcash * 100
            print(f"策略结束：股票 {data._name} 最终市值: {self.broker.get_value():.2f}, 利润: {profit:.2f}, 盈利率: {profit_per:.2f}%")

# 获取股票数据并转换时间
def convert_trade_date(df):
    # 将 'trade_date' 列从 Unix 时间戳转换为 pandas datetime 类型
    if df['trade_date'].dtype == 'O':
        df['trade_date'] = pd.to_datetime(df['trade_date'])

    # 确保数据按日期升序排序
    df = df.sort_values(by='trade_date')

    return df

# 使用pandas导出回测结果
def export_to_excel_openpyxl_with_mapping(
    data_list, filename="profit_loss_report.xlsx", field_mapping=None, field_order=None
):
    """
    使用 pandas + openpyxl 导出 Excel，支持字段映射、顺序和盈亏列标红/绿
    
    Args:
        data_list: 字典列表数据
        filename: 导出文件名
        field_mapping: 字段映射字典
        field_order: 字段顺序列表
    """
    if not data_list:
        print("数据列表为空，无法导出")
        return False

    try:
        # 1. 创建 DataFrame
        df = pd.DataFrame(data_list)

        # 2. 按字段顺序重排
        if field_order:
            available_fields = [f for f in field_order if f in df.columns]
            df = df[available_fields]

        # 3. 对数据进行排序
        df = df.sort_values(by=["no", "holding_days", "stock_code", "stock_name", "bought_date"], ascending=[True, True, True, True, True])

        # 3. 映射列名
        if field_mapping:
            rename_dict = {k: v for k, v in field_mapping.items() if k in df.columns}
            df = df.rename(columns=rename_dict)

        # 4. 导出到 Excel
        df.to_excel(filename, index=False)

        # 5. 用 openpyxl 打开并添加颜色
        wb = load_workbook(filename)
        ws = wb.active

        # 红/绿填充
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

        # 找到需要标色的列
        color_cols = ["市值(元)"]
        color_col_idx = {}
        for idx, col in enumerate(ws[1], start=1):
            if col.value in color_cols:
                color_col_idx[col.value] = idx

        # 从第二行开始循环加颜色
        for row in range(2, ws.max_row + 1):
            for col_name, col_idx in color_col_idx.items():
                cell_market = ws.cell(row=row, column=col_idx)
                cell_profit = ws.cell(row=row, column=col_idx+1)
                cell_profit_percent = ws.cell(row=row, column=col_idx+2)
                cell_profit_percent.value = str(f"{float(cell_profit_percent.value):.2f}%")
                cell_market.number_format = "0.00"
                cell_profit.number_format = "0.00"
                try:
                    cell_profit_value = float(cell_profit.value) if cell_profit.value is not None else 0
                    if cell_profit_value > 0:
                        cell_market.fill = red_fill  # 正收益红色
                        cell_profit.fill = red_fill
                        cell_profit_percent.fill = red_fill
                    elif cell_profit_value < 0:
                        cell_market.fill = green_fill    # 负收益绿色
                        cell_profit.fill = green_fill
                        cell_profit_percent.fill = green_fill
                    # 0 或 None 不着色（或可自定义）
                except (ValueError, TypeError) as e:
                    # print(f"Row {row}, Column {col_name}: Invalid value {cell_market.value}, Error: {e}")
                    continue
        
        cell_summary1 = ws.cell(row=ws.max_row, column=color_col_idx["市值(元)"]+1)
        # 设置居中（水平+垂直）
        cell_summary1.alignment = Alignment(horizontal="center", vertical="center")
        # 设置加粗
        cell_summary1.font = Font(bold=True)

        cell_summary2 = ws.cell(row=ws.max_row, column=color_col_idx["市值(元)"]-1)
        # 设置居中（水平+垂直）
        cell_summary2.alignment = Alignment(horizontal="center", vertical="center")
        # 设置加粗
        cell_summary2.font = Font(bold=True)

        cell_summary3 = ws.cell(row=ws.max_row, column=color_col_idx["市值(元)"]-4)
        # 设置居中（水平+垂直）
        cell_summary3.alignment = Alignment(horizontal="center", vertical="center")
        # 设置加粗
        cell_summary3.font = Font(bold=True)

        wb.save(filename)

        print(f"✅ 回测数据已成功导出到 {filename}")
        # print(f"📖 回测数据预览:")
        # print(df.head())

        return True

    except Exception as e:
        print(f"❌ 导出Excel文件时出错: {e}")
        return False

# 导出盈亏报告
def export_profit_loss_report_excel(data_list, holding_day, filename=None, add_timestamp=True):
    """导出盈亏报告 Excel 版本"""
    if filename is None:
        base_name = "组合盈亏报告"
    else:
        if filename.endswith('.xlsx'):
            base_name = filename[:-5]
        else:
            base_name = filename

    if add_timestamp:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"持有{holding_day}天{base_name}_{timestamp}.xlsx"
    else:
        filename = f"持有{holding_day}天{base_name}.xlsx"

    return export_to_excel_openpyxl_with_mapping(
        data_list=data_list,
        filename=filename,
        field_mapping=PROFIT_LOSS_MAPPING,
        field_order=PROFIT_LOSS_ORDER
    )

# 获取 df 中指定 trade_date 后的第n条记录的 trade_date 值
def get_nth_next_trade_date(df, trade_date, stock_code, n):
    """
    获取 df 中指定 trade_date 后的第2条记录的 trade_date 值。
    
    参数：
        df: DataFrame，索引为 trade_date，包含 adj_close_price 等列
        trade_date: 目标交易日（datetime 或字符串）
        stock_code: 股票代码（用于日志）
    
    返回：
        datetime 或 None：后第2条记录的 trade_date
    """
    # 确保 trade_date 为 datetime
    if isinstance(trade_date, str):
        trade_date = pd.to_datetime(trade_date)
    
    # 检查 trade_date 是否在索引中
    if trade_date not in df.index:
        print(f"股票 {stock_code} 在 {trade_date} 无记录")
        return None
    
    # 获取 trade_date 的索引位置
    idx_position = df.index.get_loc(trade_date)
    
    # 检查是否有足够的后续记录（需要至少2条）
    if idx_position + n >= len(df):
        print(f"股票 {stock_code} 在 {trade_date} 后记录不足2条")
        return None
    
    # 返回后第2条记录的 trade_date
    return df.index[idx_position + n]

# 在交易的第一天，如果满足条件，则添加一条第一支撑位的交易数据，以支持交易当天回踩时买入
def insert_first_support_price_data(df, first_support_price):
    
    # 获取第 0 行数据
    high_value = df.iloc[0]['high']
    low_value = df.iloc[0]['low']
    open_value = df.iloc[0]['open']
    close_value = df.iloc[0]['close']
    current_date = pd.to_datetime(df.iloc[0]['trade_date'])
    industry_level2 = df.iloc[0]['industry_level2']
    industry_level3 = df.iloc[0]['industry_level3']
    next_2_hour = current_date + timedelta(hours=2)
    
    # 计算新数据的 low 和 close
    new_open_value = (first_support_price + open_value) / 2
    new_value = (open_value + close_value) / 2
    
    # 创建新数据
    if first_support_price >= low_value and first_support_price <= high_value and new_open_value < high_value:
        new_data = pd.DataFrame({
            'stock_code': [df.iloc[0]['stock_code']],
            'trade_date': [next_2_hour],
            'stock_name': [df.iloc[0]['stock_name']],
            'volume': [df.iloc[0]['volume']],
            'close': [close_value],     # 保持原 close
            'high': [high_value],       # 保持原 high
            'low': [low_value],         # 保持原 low
            'open': [new_open_value],   # 使用 new_close_value 作为 open
            'industry_level2': [industry_level2],
            'industry_level3': [industry_level3]
        })
    else:
        new_data = pd.DataFrame({
            'stock_code': [df.iloc[0]['stock_code']],
            'trade_date': [next_2_hour],
            'stock_name': [df.iloc[0]['stock_name']],
            'volume': [df.iloc[0]['volume']],
            'close': [close_value],     # 保持原 close
            'high': [high_value],       # 保持原 high
            'low': [low_value],         # 保持原 low
            'open': [new_value],        # 使用 new_value 作为 open
            'industry_level2': [industry_level2],
            'industry_level3': [industry_level3]
        })
    
    # 合并数据
    df = pd.concat([df, new_data], ignore_index=True)
    
    df = convert_trade_date(df)
    
    return df

# 定义回测流程，支持多股票批量回测
def run_backtest(stock_data_list, history_trading_days, holding_days, total_initial_cash):
    global PROFIT_AND_LOSS_SITUATION, INITIAL_CASH

    INITIAL_CASH = total_initial_cash / len(stock_data_list)

    # 运行之前，先清空全局变量
    if len(PROFIT_AND_LOSS_SITUATION) > 0:
        PROFIT_AND_LOSS_SITUATION.clear()

    # 遍历每个突破日股票，分别加载数据和添加策略
    for holding_day in holding_days:
        # 清空结果
        PROFIT_AND_LOSS_SITUATION.clear()
        for index, record in enumerate(stock_data_list):
            cerebro = bt.Cerebro()
            # 设置初始现金
            cerebro.broker.set_cash(INITIAL_CASH)
            # 设置佣金
            cerebro.broker.setcommission(commission=0.001)
            # 设置滑点
            cerebro.broker.set_slippage_perc(0.001)

            stock_code = record['stock_code']
            trade_date = record['trade_date']
            stock_name = record['stock_name']
            first_support_price = round(record['first_support_price'], 2)

            # 获取突破日后N日行情数据（单只股票）
            df = get_next_N_days_data([record], holding_day)

            # 调整一下第一天的交易数据，使可以支持回踩买入
            df = insert_first_support_price_data(df, first_support_price)
            if df.empty:
                print(f"股票 {stock_code} 在 {trade_date} 后无数据，跳过")
                continue

            # 转换时间
            df = convert_trade_date(df)
            df.set_index('trade_date', inplace=True)
            df.sort_index(inplace=True)

            # ==== 数据清洗 ====
            df.replace([np.inf, -np.inf], np.nan, inplace=True)
            df.dropna(subset=['close', 'open', 'volume'], inplace=True)
            df = df[df['volume'] >= 0]

            if df.empty:
                print(f"股票{stock_code}数据清洗后为空，跳过")
                continue

            # 检查数据完整性
            required_columns = ['close', 'open', 'high', 'low', 'volume']
            if not all(col in df.columns for col in required_columns):
                print(f"股票 {stock_code} 缺少必要列，跳过")
                continue

            # 自定义Pandas数据类，包含复权价
            class PandasData(bt.feeds.PandasData):
                lines = ('open', 'high', 'low', 'close', 'volume', 'openinterest')
                params = (
                    ('datetime', None),  # 使用索引作为时间
                    ('open', 'open'),
                    ('high', 'high'),
                    ('low', 'low'),
                    ('close', 'close'),
                    ('volume', 'volume'),
                    ('openinterest', None),  # 股票数据无未平仓合约，设为None
                )

            data_feed = PandasData(dataname=df)
            # print(f"股票 {stock_code} 数据预览:\n{data_feed._dataname.head()}")

            # 加载该股票数据
            try:
                cerebro.adddata(data_feed, name=stock_code)
            except Exception as e:
                print(f"加载股票 {stock_code} 数据失败: {e}")
                continue

            # 添加策略时传入对应参数
            cerebro.addstrategy(MyStrategy,
                breakout_date=trade_date,
                first_support_price=first_support_price)

            start_trade_date = df.iloc[0].name
            start_trade_date_str = start_trade_date.date().strftime('%Y-%m-%d')
            current_trade_date = get_nth_next_trade_date(df, start_trade_date, stock_code, holding_day)
            current_trade_date_str = current_trade_date.date().strftime('%Y-%m-%d')
            print(f"=======================开始对股票[{stock_name}]模拟在[{start_trade_date_str}]日买入时持有[{holding_day}]天的回测======================")
            print(f"开始市值: {cerebro.broker.startingcash:.2f}\n")
            begin_value = cerebro.broker.startingcash
            cerebro.run()
            end_value = cerebro.broker.getvalue()
            profit = end_value - begin_value
            profit_percent = profit / begin_value * 100
            PROFIT_AND_LOSS_SITUATION.append({
                "no": index,
                "stock_code": stock_code,
                "stock_name": stock_name,
                "bought_date": start_trade_date_str,
                "init_cash": INITIAL_CASH,
                "holding_days": holding_day,
                "trade_date": current_trade_date_str,
                "market_value": end_value,
                "profit": profit,
                "profit_percent": profit_percent
            })
            print(f"结束市值: {cerebro.broker.getvalue():.2f}\n")
            print(f"=======================结束对股票[{stock_name}]模拟在[{start_trade_date_str}]日买入时持有[{holding_day}]天的回测======================")
            print("\n")

        total_market_value = sum(item.get("market_value", 0) for item in PROFIT_AND_LOSS_SITUATION)
        PROFIT_AND_LOSS_SITUATION.append({
            "no":max(item.get("no", 0) for item in PROFIT_AND_LOSS_SITUATION)+1,
            "stock_code": "",
            "stock_name": "",
            "bought_date": "初始金额",
            "init_cash": total_initial_cash,
            "holding_days": "",
            "trade_date": "结算市值",
            "market_value": total_market_value,
            "profit": "组合盈利率:",
            "profit_percent": (total_market_value - total_initial_cash) / total_initial_cash * 100
        })
        export_profit_loss_report_excel(PROFIT_AND_LOSS_SITUATION, holding_day)


if __name__ == "__main__":

    # 创建 ConfigParser 对象
    config = configparser.ConfigParser()

    # 读取 .conf 文件
    config.read('./config.conf')
    cond1_and_cond3=config['settings']['cond1_and_cond3']                                           # 条件1和条件3的配置项。
    total_initial_cash_settings=config['settings']['total_initial_cash']                            # 初始金额
    holdingdays_settings=config['settings']['holdingdays']                                          # 持有天数配置
    history_trading_days=cond1_and_cond3.split('_')[0]
    holding_days = [int(x.strip()) for x in holdingdays_settings.split(',')]
    total_initial_cash = float(total_initial_cash_settings)

    # 读取 CSV 文件，跳过第一行作为列名称
    df = pd.read_csv('stock_query_results_20250813_122711_cond1.1_40days_25per_35per_no_cond2_yes_cond5.csv', header=0)

    # 将 DataFrame 转换为指定格式的列表
    stock_data_list = df[['交易日期', '股票代码', '股票名称', '前复权_收盘价', '前复权_前N天最高收盘价']].to_dict('records')

    # 重命名键为 trade_date 和 stock_code
    for d in stock_data_list:
        d['trade_date'] = d.pop('交易日期')
        d['stock_code'] = d.pop('股票代码')
        d['stock_name'] = d.pop('股票名称')
        d['adj_close_price'] = d.pop('前复权_收盘价')
        d['first_support_price'] = d.pop('前复权_前N天最高收盘价')

    # 执行批量回测
    run_backtest(stock_data_list, history_trading_days, holding_days, total_initial_cash)
